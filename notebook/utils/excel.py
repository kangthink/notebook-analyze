import os

from openpyxl import load_workbook

class ExcelContainer:
    
    def __init__(self, path):
        self._file_root = path[:-len(os.path.basename(path))]
        self._file_name = os.path.basename(path)
        self._workbook = load_workbook(filename=path)
        
    def worksheet(self, name):
        try:
            return self._workbook[name]
        except KeyError as e:
            return Exception
    
    @staticmethod
    def header(worksheet):
        header = [cell.value for row in worksheet.iter_rows(min_row=1, max_row=1) for cell in row]
        return header
    
    @staticmethod
    def elements(worksheet, start=2):
        for row in worksheet.iter_rows(min_row=start):
            values = [cell.value for cell in row]
            yield values